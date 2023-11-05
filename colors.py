from openpyxl.styles import PatternFill

# Define custom fill colors for each month
MONTHS_COLORS = [
    PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid"),   # Blue (January)
    PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),   # Orange (February)
    PatternFill(start_color="800080", end_color="800080", fill_type="solid"),   # Purple (March)
    PatternFill(start_color="008000", end_color="008000", fill_type="solid"),   # Green (April)
    PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid"),   # Gold (May)
    PatternFill(start_color="4B0082", end_color="4B0082", fill_type="solid"),   # Indigo (June)
    PatternFill(start_color="FF4500", end_color="FF4500", fill_type="solid"),   # OrangeRed (July)
    PatternFill(start_color="00FFFF", end_color="00FFFF", fill_type="solid"),   # Cyan (August)
    PatternFill(start_color="800000", end_color="800000", fill_type="solid"),   # Maroon (September)
    PatternFill(start_color="2E8B57", end_color="2E8B57", fill_type="solid"),   # SeaGreen (October)
    PatternFill(start_color="000080", end_color="000080", fill_type="solid"),   # Navy (November)
    PatternFill(start_color="8B4513", end_color="8B4513", fill_type="solid")    # SaddleBrown (December)
]

def color_of_month(date_in_month):
    month = date_in_month.month
    return MONTHS_COLORS[month -1]

def color_of_weekend():
    # Grey fill
    return PatternFill(start_color="808080", end_color="808080", fill_type="solid")

def color_of_holiday():
    # Red fill
    return PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

def color_of_vaccation():
   # Green fill
    return PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

def color_of_approved_vaccation():
    # Dark green fill
    return PatternFill(start_color="006400", end_color="006400", fill_type="solid")

def color_of_others():
    # Yellow fill
    return PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

def color_of_parental_leave():
    # Purple fill
    return PatternFill(start_color="CC99FF", end_color="CC99FF", fill_type="solid")

def color_of_education():
    #himmel_farg_fill
    return PatternFill(start_color="00CCFF", end_color="00CCFF", fill_type="solid")