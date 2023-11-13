import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image
import argparse

from copy import copy
import date_utils as date_utils
import colors as colors
import red_days as red_days
import utils

################### DO NOT CHANGE!!! ##################
EXCEL_MONTH_ROW = 2
EXCEL_WEEKNUMBER_ROW = 3
EXCEL_DATE_ROW = 4
EXCEL_WEEK_DATE_ROW = 5

EXCEL_FIRST_EMPLOYEE_ROW = 6
EXCEL_LAST_EMPLOYEE_ROW = 123
EXCEL_SOURCE_SHEET = 'Template-sheet'
EXCEL_FIRST_VACATION_COLUMN_LETTER = 'E'

######################################################

################### User preferences ###################
global custom_password
global custom_year
#########################################################

def calculate_column_letter(start_date, date_value):
    date_count = date_utils.number_dates_in_period(start_date, date_value)
    column_index = openpyxl.utils.column_index_from_string(EXCEL_FIRST_VACATION_COLUMN_LETTER) + date_count -1
    return get_column_letter(column_index)


def clone_sheet(source_wb, destination_wb, source_sheet_name, new_sheet_name):
    workbook = openpyxl.load_workbook(source_wb)
    source_sheet = workbook[source_sheet_name]
    new_sheet = workbook.copy_worksheet(source_sheet)
    new_sheet.title = new_sheet_name
    workbook.save(destination_wb)


def set_start_date_of_period(workbook_path, sheet_name, start_date):
    wb = openpyxl.load_workbook(workbook_path)
    ws = wb[sheet_name]
    ws[f'{EXCEL_FIRST_VACATION_COLUMN_LETTER}{EXCEL_DATE_ROW}'].value = start_date
    ws[f'{EXCEL_FIRST_VACATION_COLUMN_LETTER}{EXCEL_WEEK_DATE_ROW}'].value = date_utils.get_weekday_short_text(start_date)
    wb.save(workbook_path)


def set_date_on_column(destination_column, source_column, date_diff):
    destination_column[EXCEL_DATE_ROW - 1].value = date_utils.date_add(source_column[EXCEL_DATE_ROW - 1].value, days=date_diff)
    destination_column[EXCEL_WEEK_DATE_ROW -1].value = date_utils.get_weekday_short_text(destination_column[EXCEL_DATE_ROW - 1].value)
    

def copy_format_column(ws, to_col_letter, date_diff):
    source_column = ws[EXCEL_FIRST_VACATION_COLUMN_LETTER]
    destination_column = ws[to_col_letter]

    # Copy the format from the source column to the newly inserted column
    for source_cell, new_cell in zip(source_column, destination_column):
        new_cell.font = copy(source_cell.font)
        new_cell.fill = copy(source_cell.fill)
        new_cell.border = copy(source_cell.border)
        new_cell.alignment = copy(source_cell.alignment)
        new_cell.number_format = source_cell.number_format
        new_cell.protection = copy(source_cell.protection)
    
    # Set the width of the newly inserted column to match the starting column
    ws.column_dimensions[to_col_letter].width = ws.column_dimensions[EXCEL_FIRST_VACATION_COLUMN_LETTER].width
    
    # Copy date formulas to display "day" on row 4th and "weekday" on row 5th
    set_date_on_column(destination_column, source_column, date_diff)

def fill_period_with_columns(workbook_path, sheet_name, date_count):
    workbook = openpyxl.load_workbook(workbook_path)
    worksheet = workbook[sheet_name]
    
    from_col_idx = openpyxl.utils.column_index_from_string(EXCEL_FIRST_VACATION_COLUMN_LETTER)
    
    # Insert columns to the right
    worksheet.insert_cols(idx=from_col_idx + 1, amount=date_count)

    for i in range(1, date_count):
        to_col_letter = get_column_letter(from_col_idx + i)
        copy_format_column(worksheet, to_col_letter, i)

    # Save the updated Excel workbook
    workbook.save(workbook_path)

def colourize_cells_to_last_employees(worksheet, start_column_index, 
                                      day_diff, color_fill, locked, 
                                      comment, image_name):
    for row in range(EXCEL_DATE_ROW, EXCEL_LAST_EMPLOYEE_ROW + 1):
        cell = worksheet.cell(row=row, column=start_column_index + day_diff)
        cell.fill = color_fill
        if row == EXCEL_DATE_ROW:
            cell.comment = comment
        
        # Cells at and above EXCEL_WEEK_DATE_ROW are always protected
        if row > EXCEL_WEEK_DATE_ROW:
            cell.protection = openpyxl.styles.protection.Protection(locked=locked)
        else:
            cell.protection = openpyxl.styles.protection.Protection(locked=True)

    if image_name:
        image = Image(f"images/{image_name}")
        width = image.width
        height = image.height
        letter = get_column_letter(start_column_index + day_diff)
        worksheet.add_image(image, f"{letter}1")
        image = worksheet._images[len(worksheet._images) - 1]

        # Justera storleken på bilden om så önskas
        image.width = int(width / 3)
        image.height = int(height / 3)

def colourize_and_lock_weekend(workbook_path, sheet_name, start_date, date_count):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(workbook_path)
    
    # Select the worksheet
    worksheet = workbook[sheet_name]
    
    # Get the index of the starting column
    start_column_index = openpyxl.utils.column_index_from_string(EXCEL_FIRST_VACATION_COLUMN_LETTER)

    heldagar = red_days.get_heldagar_as_dates(custom_year)
    kortdagar = red_days.get_kortagar_as_dates(custom_year)

    for date_diff in range(0, date_count):
        current_date = date_utils.date_add(input_date_obj=start_date, days=date_diff)
        color_fill = PatternFill(fill_type=None)
        locked = False
        if date_utils.is_Weekend(current_date):
            color_fill = colors.color_of_weekend()
            locked = True
        
        heldag_name = red_days.get_heldag_name(current_date, heldagar)
        kortdag_name = red_days.get_kortdag_name(current_date, kortdagar)
        
        comment = None
        image_name = None
        if heldag_name:
            color_fill = colors.color_of_heldag()
            locked = True
            comment = Comment(heldag_name, None)
            image_name = red_days.get_image_path(heldag_name)

        if kortdag_name:
            color_fill = colors.color_of_kortdag()
            comment = Comment(kortdag_name, None)
            image_name = red_days.get_image_path(kortdag_name)

        colourize_cells_to_last_employees(worksheet, start_column_index, 
                                          date_diff, color_fill, locked,
                                            comment, image_name)

    # Save the updated Excel workbook
    workbook.save(workbook_path)

def fill_month_headers(workbook_path, sheet_name, start_date, end_date):
    wb = openpyxl.load_workbook(workbook_path)
    ws = wb[sheet_name]
    
    # Get the index of the starting column
    monthly_ranges = date_utils.generate_monthly_ranges_within_year(start_date, end_date)    
    month = 0
    for (month_start, month_end) in monthly_ranges:
        if month_start and month_end:
            month_start_column_letter = calculate_column_letter(start_date, month_start)
            month_end_column_letter = calculate_column_letter(start_date, month_end)
            start_cell_ref = f"{month_start_column_letter}{EXCEL_MONTH_ROW}"
            end_cell_ref = f"{month_end_column_letter}{EXCEL_MONTH_ROW}"
            merge_range = f"{start_cell_ref}:{end_cell_ref}"
            ws.merge_cells(merge_range)
            merged_cell = ws[f"{start_cell_ref}"]
            # Set the month as string in the merged cells
            merged_cell.value = date_utils.MONTHS_OF_A_YEAR[month]

            # Colourize the month
            color_to_fill = colors.MONTHS_COLORS[month]
            merged_cell.fill = color_to_fill
        month = month + 1

    # Save the updated Excel workbook
    wb.save(workbook_path)

def fill_week_headers(workbook_path, sheet_name, start_date, end_date):
    workbook = openpyxl.load_workbook(workbook_path)
    worksheet = workbook[sheet_name]
    
    week_ranges = date_utils.week_ranges_in_range(start_date, end_date)

    for (week_number, (week_start, week_end)) in week_ranges:
        # Find cells to merge aka making a week
        week_start_column_letter = calculate_column_letter(start_date, week_start)
        week_end_column_letter = calculate_column_letter(start_date, week_end)
        start_cell_ref = f"{week_start_column_letter}{EXCEL_WEEKNUMBER_ROW}"
        end_cell_ref = f"{week_end_column_letter}{EXCEL_WEEKNUMBER_ROW}"

        worksheet.merge_cells(f"{start_cell_ref}:{end_cell_ref}")
        
        # Set week number
        merged_cell = worksheet[start_cell_ref]
        merged_cell.value = f"Week {week_number}"

    # Save the updated Excel workbook
    workbook.save(workbook_path)


def apply_conditional_formatting(workbook_path, sheet_name, range_to_format):
    workbook = openpyxl.load_workbook(workbook_path)
    worksheet = workbook[sheet_name]

    vaccation_rule = CellIsRule(operator='equal', formula=['"v"'], 
                                stopIfTrue=True, fill=colors.color_of_vaccation())
    approved_vaccation_rule = CellIsRule(operator='equal', formula=['"a"'], 
                                         stopIfTrue=True, fill=colors.color_of_approved_vaccation())
    other_rule = CellIsRule(operator='equal', formula=['"o"'], 
                            stopIfTrue=True, fill=colors.color_of_others())
    parental_rule = CellIsRule(operator='equal', formula=['"p"'], 
                               stopIfTrue=True, fill=colors.color_of_parental_leave())
    education_rule = CellIsRule(operator='equal', formula=['"e"'], 
                                stopIfTrue=True, fill=colors.color_of_education())

    # Add the rules
    worksheet.conditional_formatting.add(range_to_format, vaccation_rule)
    worksheet.conditional_formatting.add(range_to_format, approved_vaccation_rule)
    worksheet.conditional_formatting.add(range_to_format, other_rule)
    worksheet.conditional_formatting.add(range_to_format, parental_rule)
    worksheet.conditional_formatting.add(range_to_format, education_rule)

    workbook.save(workbook_path)


def create_vaccation_period(source_wb, destination_wb, sheet_name, start_date, end_date):
    clone_sheet(source_wb, destination_wb, EXCEL_SOURCE_SHEET, sheet_name)
    set_start_date_of_period(destination_wb, sheet_name, start_date)

    date_count = date_utils.number_dates_in_period(start_date, end_date)

    # Insert columns to the right with the same format as start_column_letter
    fill_period_with_columns(destination_wb, sheet_name, date_count)
    colourize_and_lock_weekend(destination_wb, sheet_name, start_date, date_count)
    fill_month_headers(destination_wb,sheet_name,start_date,end_date)
    fill_week_headers(destination_wb,sheet_name,start_date,end_date)

    # Caculate range_format
    end_column_index = openpyxl.utils.column_index_from_string(EXCEL_FIRST_VACATION_COLUMN_LETTER) + date_count -1
    end_column_letter = get_column_letter(end_column_index)

    apply_conditional_formatting(destination_wb, sheet_name, 
                                 f'{EXCEL_FIRST_VACATION_COLUMN_LETTER}{EXCEL_FIRST_EMPLOYEE_ROW}:{end_column_letter}{EXCEL_LAST_EMPLOYEE_ROW}')

    workbook = openpyxl.load_workbook(destination_wb)
    worksheet = workbook[sheet_name]
    # Auto filter on B5:D5
    worksheet.auto_filter.ref = "B5:D5"
    # Protect the sheet with a password
    worksheet.protection.sheet = True
    worksheet.protection.password = custom_password
    # Move up
    workbook.move_sheet(worksheet, -2)
    worksheet["B1"].value = f"Vaccation list {custom_year}"
    worksheet["B3"].value = f"{sheet_name}"
    workbook.save(destination_wb)

def add_vacation_as_new_sheet(source_wb, destination_wb, start_month, end_month):
    start_date = date_utils.first_day_of_month(custom_year, start_month)
    end_date=date_utils.last_day_of_month(custom_year, end_month)
    create_vaccation_period(source_wb=source_wb, 
                                destination_wb=destination_wb, 
                                sheet_name=
                                f"{date_utils.text_of_month(start_date)}-{date_utils.text_of_month(end_date)}",
                                start_date=start_date, end_date=end_date)
def main(custom_periods):
    source_wb='templates/vaccation-template.xlsx'
    destination_wb=f'output/{custom_year}-vaccation.xlsx'
    
    for start_month, end_month in custom_periods:
        add_vacation_as_new_sheet(source_wb, destination_wb, start_month, end_month)
        source_wb = destination_wb
        
    # Remove the template-sheet
    workbook = openpyxl.load_workbook(destination_wb)
    template_sheet = workbook[EXCEL_SOURCE_SHEET]
    workbook.remove(template_sheet)
    properties = workbook.properties
    properties.creator = "Thành Trương"  # Set your name as the author
    properties.description = "Created by Python program written by Thành Trương - 2024"
    

    workbook.save(destination_wb)

parser = argparse.ArgumentParser()
parser.add_argument("year", type=int, help="Year")
parser.add_argument("--password", type=str, default='12345', help="Password. Default is 12345 ")
parser.add_argument("--periods", type=str, default='[1- 4,5 -8, 9 - 12]', 
                    help="How a year is split into periods. \
                          The default is [1-4, 5-8, 9-12 ] which means the vacation list has 3 sheets: \
                            January-April, May-August, and September-Decemnber")

if __name__ == "__main__":
    args = parser.parse_args()
    custom_year = args.year
    custom_password = args.password
    input_periods = args.periods
    print(f"{custom_year} {custom_password} {input_periods}")
    
    main(custom_periods=utils.parse_input_periods(input_periods))
    

