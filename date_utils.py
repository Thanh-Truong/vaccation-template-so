
from datetime import datetime, timedelta
from datetime import date

MONTHS_OF_A_YEAR = [
    "January",
    "February",
    "March",
    "April",
    "May",
    "June",
    "July",
    "August",
    "September",
    "October",
    "November",
    "December"
]

from openpyxl.styles import PatternFill

# Define custom fill colors for each month
MONTHS_COLORS = [
    PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid"),   # Blue (January)
    PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),   # Orange (February)
    PatternFill(start_color="800080", end_color="800080", fill_type="solid"),   # Purple (March)
    PatternFill(start_color="008000", end_color="008000", fill_type="solid"),   # Green (April)
    PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid"),   # Gold (May)
    PatternFill(start_color="4B0082", end_color="4B0082", fill_type="solid"),   # Indigo (June)
    PatternFill(start_color="FF4500", end_color="FF4500", fill_type="solid"),   # OrangeRed (July)
    PatternFill(start_color="00FFFF", end_color="00FFFF", fill_type="solid"),   # Cyan (August)
    PatternFill(start_color="800000", end_color="800000", fill_type="solid"),   # Maroon (September)
    PatternFill(start_color="2E8B57", end_color="2E8B57", fill_type="solid"),   # SeaGreen (October)
    PatternFill(start_color="000080", end_color="000080", fill_type="solid"),   # Navy (November)
    PatternFill(start_color="8B4513", end_color="8B4513", fill_type="solid")    # SaddleBrown (December)
]

def color_of_month(date_in_month):
    month = date_in_month.month
    return MONTHS_COLORS[month -1]

def text_of_month(date_in_month):
    month = date_in_month.month
    return MONTHS_OF_A_YEAR[month -1]

def count_dates_between_inclusive(start_date, end_date):
    # Calculate the number of days between the two dates
    days_between = (end_date - start_date).days

    # Return the count of dates between (inclusive)
    return max(0, days_between + 1)

def parse_date(date_string, date_format='%Y-%m-%d'):
    try:
        date_obj = datetime.strptime(date_string, date_format)
        return date_obj.date
    except ValueError:
        return None

def date_add(input_date_obj, days=0, months=0, years=0):
    # Create a timedelta to add the specified number of days, months, and years
    delta = timedelta(days=days)

    # Add the timedelta to the input date
    return input_date_obj + delta


def get_weekday(date_obj):
    # Get the weekday as an integer (0 = Monday, 1 = Tuesday, ..., 6 = Sunday)
    return date_obj.weekday()

def is_Weekend(date_obj):
    return get_weekday(date_obj) == 5 or get_weekday(date_obj) == 6

def end_of_month(year, month):
    # Calculate the last day of the next month
    if month == 12:
        next_month = 1
        next_year = year + 1
    else:
        next_month = month + 1
        next_year = year

    # Create a datetime object for the first day of the next month
    first_day_of_next_month = date(next_year, next_month, 1)

    # Subtract one day from the first day of the next month to get the last day of the specified month
    last_day_of_month = first_day_of_next_month - timedelta(days=1)

    return last_day_of_month

from datetime import date, timedelta

def add_months(sourcedate, months):
    # Calculate the new date by adding the specified number of months
    month = sourcedate.month - 1 + months
    year = sourcedate.year + month // 12
    month = month % 12 + 1
    day = min(sourcedate.day, (date(year, month + 1, 1) - timedelta(days=1)).day)
    return date(year, month, day)

def find_interval_intersection(start_date, end_date, first_day_of_month, last_day_of_month):
    # Ensure that the intervals are valid
    if start_date > end_date or first_day_of_month > last_day_of_month:
        return None  # Invalid intervals

    # Determine the intersection start date and end date
    intersection_start = max(start_date, first_day_of_month)
    intersection_end = min(end_date, last_day_of_month)

    # Check if there is a valid intersection
    if intersection_start <= intersection_end:
        return (intersection_start, intersection_end)
    else:
        return (None, None)  # No intersection
    
def generate_monthly_ranges_within_year(start_date, end_date):
    monthly_ranges = []

    # Ensure the provided start_date is in the same year as end_date
    if start_date.year != end_date.year:
        raise ValueError("start_date and end_date should be in the same calendar year")

    current_date = date(start_date.year, 1, 1)

    for month in range(1, 13):
        # Calculate the first day of the month
        first_day_of_month = date(current_date.year, month, 1)

        # Calculate the last day of the month
        if month == 12:
            last_day_of_month = date(current_date.year, 12, 31)
        else:
            last_day_of_month = (date(current_date.year, month + 1, 1) - timedelta(days=1))

        # Check if the calculated month is within the given range
        monthly_ranges.append(find_interval_intersection(start_date, end_date, 
                                                         first_day_of_month, last_day_of_month))

    return monthly_ranges

def week_ranges_in_range(start_date, end_date):
    if start_date.year != end_date.year:
        raise ValueError("start_date and end_date should be in the same calendar year")

    week_ranges = []
    current_date = start_date

    while current_date <= end_date:
        # Calculate the ISO week number for the current date
        week_number = current_date.isocalendar()[1]

        # Calculate the first day of the week (Monday)
        first_day_of_week = current_date - timedelta(days=current_date.weekday())

        # Calculate the end day of the week (Sunday)
        end_day_of_week = first_day_of_week + timedelta(days=6)
        (star, end) = find_interval_intersection(start_date, end_date, first_day_of_week, end_day_of_week)
        week_ranges.append((week_number, (star, end)))

        current_date += timedelta(days=7)

    return week_ranges

if __name__ == "__main__":
    # Example usage:
    start_date = date(2024, 1, 1)
    
    end_date = date(2024, 1, 10)
    #result = generate_monthly_ranges_within_year(start_date, end_date)
    #print(result)

    week_numbers = week_ranges_in_range(start_date, end_date)
    for (week_number, (start, end)) in week_numbers:
        print(f"{start} {end}")