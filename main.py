import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.styles import Color
from openpyxl.formatting.rule import CellIsRule
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation


from copy import copy
import date_utils
import colors
from datetime import date
import red_days

################### DO NOT CHANGE!!! ##################
EXCEL_MONTH_ROW = 2
EXCEL_WEEKNUMBER_ROW = 3
EXCEL_DATE_ROW = 4
EXCEL_WEEK_DATE_ROW = 5

EXCEL_FIRST_EMPLOYEE_ROW = 6
EXCEL_LAST_EMPLOYEE_ROW = 123
######################################################

################### User preferences ###################
CUSTOM_PASSWORD = '12345'
CUSTOM_YEAR = 2024
#########################################################

def date_to_column_letter(start_column_letter, start_date, a_date):
    diff = date_utils.count_dates_between_inclusive(start_date, a_date)
    column_index = openpyxl.utils.column_index_from_string(start_column_letter) + diff -1
    return get_column_letter(column_index)

# source_wb, 
# destination_wb, 
# new_sheet_name
def clone_a_sheet(source_wb, destination_wb, new_sheet_name):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(source_wb)

    # Select the sheet you want to clone
    source_sheet = workbook['Template-sheet']

    # Create a new sheet to clone into
    new_sheet = workbook.copy_worksheet(source_sheet)
    new_sheet.title = new_sheet_name  # Set a new name for the cloned sheet

    # Save the updated Excel workbook
    workbook.save(destination_wb)

def set_start_date(workbook_path, sheet_name, start_column_letter, start_date):
    # Load the Excel workbook
    wb = openpyxl.load_workbook(workbook_path)

    # Select the sheet
    ws = wb[sheet_name]

    # Create a new sheet to clone into
    ws[f'{start_column_letter}{EXCEL_DATE_ROW}'].value = start_date

    # Save the updated Excel workbook
    wb.save(workbook_path)


def copy_date_formulas(destination_column, destination_column_letter, source_column, source_column_letter, date_diff):
    # Copy formulas
    destination_column[EXCEL_DATE_ROW - 1].value = "={}{}+{}".format(source_column_letter, EXCEL_DATE_ROW, date_diff)
    destination_column[EXCEL_WEEK_DATE_ROW -1].value = '=TEXT({}{},"ddd")'.format(destination_column_letter, EXCEL_DATE_ROW)
    # Above should work but here another way to asign values instead formulas
    destination_column[EXCEL_DATE_ROW - 1].value = date_utils.date_add(source_column[EXCEL_DATE_ROW - 1].value, days=date_diff)
    destination_column[EXCEL_WEEK_DATE_ROW -1].value = date_utils.get_weekday_short_text(destination_column[EXCEL_DATE_ROW - 1].value)
    


def copy_format_column(worksheet, source_column_letter, destination_column_letter, date_diff):
    source_column = worksheet[source_column_letter]
    destination_column = worksheet[destination_column_letter]

    # Copy the format from the source column to the newly inserted column
    for source_cell, new_cell in zip(source_column, destination_column):
        new_cell.font = copy(source_cell.font)
        new_cell.fill = copy(source_cell.fill)
        new_cell.border = copy(source_cell.border)
        new_cell.alignment = copy(source_cell.alignment)
        new_cell.number_format = source_cell.number_format
        new_cell.protection = copy(source_cell.protection)
    
    # Set the width of the newly inserted column to match the starting column
    worksheet.column_dimensions[destination_column_letter].width = worksheet.column_dimensions[source_column_letter].width
    
    # Copy date formulas to display "day" on row 4th and "weekday" on row 5th
    copy_date_formulas(destination_column, destination_column_letter, 
                       source_column, source_column_letter, date_diff)

def insert_columns_with_format(workbook_path, sheet_name, source_column_letter, date_count):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(workbook_path)
    
    # Select the worksheet
    worksheet = workbook[sheet_name]
    
    # Get the index of the starting column
    start_column_index = openpyxl.utils.column_index_from_string(source_column_letter)
    
    # Insert column_count to the right 
    worksheet.insert_cols(idx=start_column_index + 1, amount=date_count)

    for date_diff in range(1, date_count):
        destination_column_letter = get_column_letter(start_column_index + date_diff)
        copy_format_column(worksheet, source_column_letter, destination_column_letter, date_diff)

    # Save the updated Excel workbook
    workbook.save(workbook_path)

def colourize_cells_to_last_employees(worksheet, start_column_index, 
                                      day_diff, color_fill, locked, comment):
    # Lock cells and fill them with GREY
    for row in range(EXCEL_DATE_ROW, EXCEL_LAST_EMPLOYEE_ROW + 1):
        cell = worksheet.cell(row=row, column=start_column_index + day_diff)
        cell.fill = color_fill
        # Cells at and above EXCEL_WEEK_DATE_ROW are always protected
        if row > EXCEL_WEEK_DATE_ROW:
            cell.protection = openpyxl.styles.protection.Protection(locked=locked)
        else:
            cell.protection = openpyxl.styles.protection.Protection(locked=True)
        cell.comment = comment

def colourize_and_lock_weekend(workbook_path, sheet_name, source_column_letter, start_date, date_count):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(workbook_path)
    
    # Select the worksheet
    worksheet = workbook[sheet_name]
    
    # Get the index of the starting column
    start_column_index = openpyxl.utils.column_index_from_string(source_column_letter)

    holidays = red_days.get_swedish_holidays_as_date_description(CUSTOM_YEAR)

    for date_diff in range(0, date_count):
        current_date = date_utils.date_add(input_date_obj=start_date, days=date_diff)
        color_fill = PatternFill(fill_type=None)
        locked = False
        if date_utils.is_Weekend(current_date):
            color_fill = colors.color_of_weekend()
            locked = True
        
        holiday_description = red_days.get_holiday_description(current_date, holidays)
        comment = None
        if holiday_description:
            color_fill = colors.color_of_holiday()
            locked = True
            column_letter = date_to_column_letter(source_column_letter, start_date, current_date)
            
            comment = Comment(holiday_description, None)

        colourize_cells_to_last_employees(worksheet, start_column_index, 
                                          date_diff, color_fill, locked, comment)

    # Save the updated Excel workbook
    workbook.save(workbook_path)

def write_months_as_headers(workbook_path, sheet_name, source_column_letter, start_date, end_date):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(workbook_path)
    
    # Select the worksheet
    worksheet = workbook[sheet_name]
    
    # Get the index of the starting column
    start_column_index = openpyxl.utils.column_index_from_string(source_column_letter)
    monthly_ranges = date_utils.generate_monthly_ranges_within_year(start_date, end_date)    
    month = 0
    for (month_start, month_end) in monthly_ranges:
        if month_start and month_end:
            # Caculate the month_start column index
            month_start_column_letter = date_to_column_letter(source_column_letter, start_date, month_start)

            # Caculate the month_end column index
            month_end_column_letter = date_to_column_letter(source_column_letter, start_date, month_end)

            # Merge a range of cells to form a month
            merge_range = f"{month_start_column_letter}{EXCEL_MONTH_ROW}:{month_end_column_letter}{EXCEL_MONTH_ROW}"
            #print("merge_range start {} end {}".format(month_start, month_end))
            #print("merge_range start column {} end column {}".format(month_start_column_letter, month_end_column_letter))
            #print(merge_range)
            worksheet.merge_cells(merge_range)
            # Get the merged cell
            merged_cell = worksheet[f"{month_start_column_letter}{EXCEL_MONTH_ROW}"]
            # Set the month as string in the merged cells
            merged_cell.value = date_utils.MONTHS_OF_A_YEAR[month]
            # Colourize the month
            color_to_fill = colors.MONTHS_COLORS[month]

            merged_cell.fill = color_to_fill

        month = month + 1

    # Save the updated Excel workbook
    workbook.save(workbook_path)

def write_weeks_as_headers(workbook_path, sheet_name, source_column_letter, start_date, end_date):
    workbook = openpyxl.load_workbook(workbook_path)
    worksheet = workbook[sheet_name]
    
    # Get the index of the starting column
    start_column_index = openpyxl.utils.column_index_from_string(source_column_letter)
    week_ranges = date_utils.week_ranges_in_range(start_date, end_date)

    for (week_number, (week_start, week_end)) in week_ranges:
        week_start_column_letter = date_to_column_letter(source_column_letter, start_date, week_start)
        week_end_column_letter = date_to_column_letter(source_column_letter, start_date, week_end)

        # Merge a range of cells to form a month
        merge_range = f"{week_start_column_letter}{EXCEL_WEEKNUMBER_ROW}:{week_end_column_letter}{EXCEL_WEEKNUMBER_ROW}"
        worksheet.merge_cells(merge_range)
        
        # Set the month as string in the merged cells
        merged_cell = worksheet[f"{week_start_column_letter}{EXCEL_WEEKNUMBER_ROW}"]
        merged_cell.value = f"Week {week_number}"

    # Save the updated Excel workbook
    workbook.save(workbook_path)


def apply_conditional_formatting(workbook_path, sheet_name, range_to_format):
    workbook = openpyxl.load_workbook(workbook_path)
    worksheet = workbook[sheet_name]

    vaccation_rule = CellIsRule(operator='equal', formula=['"v"'], 
                                stopIfTrue=True, fill=colors.color_of_vaccation())
    approved_vaccation_rule = CellIsRule(operator='equal', formula=['"a"'], 
                                         stopIfTrue=True, fill=colors.color_of_approved_vaccation())
    other_rule = CellIsRule(operator='equal', formula=['"o"'], 
                            stopIfTrue=True, fill=colors.color_of_others())
    parental_rule = CellIsRule(operator='equal', formula=['"p"'], 
                               stopIfTrue=True, fill=colors.color_of_parental_leave())
    education_rule = CellIsRule(operator='equal', formula=['"e"'], 
                                stopIfTrue=True, fill=colors.color_of_education())

    # Add the rules
    worksheet.conditional_formatting.add(range_to_format, vaccation_rule)
    worksheet.conditional_formatting.add(range_to_format, approved_vaccation_rule)
    worksheet.conditional_formatting.add(range_to_format, other_rule)
    worksheet.conditional_formatting.add(range_to_format, parental_rule)
    worksheet.conditional_formatting.add(range_to_format, education_rule)

    workbook.save(workbook_path)


def create_vaccation_period(source_wb, destination_wb, sheet_name, start_date, end_date):
    start_column_letter = 'E'

    # Clone the Template-sheet to a new sheet 
    clone_a_sheet(source_wb, destination_wb, sheet_name)
    # On the new workbook, let set the start_date at E4 
    set_start_date(destination_wb, sheet_name, start_column_letter, start_date)

    date_count = date_utils.count_dates_between_inclusive(start_date, end_date)
    #print(f"There are {date_count} inclusive dates between {start_date} and {end_date}.")

    # Insert columns to the right with the same format as start_column_letter
    insert_columns_with_format(destination_wb, sheet_name, start_column_letter, date_count)
    colourize_and_lock_weekend(destination_wb, sheet_name, start_column_letter, start_date, date_count)
    write_months_as_headers(destination_wb,sheet_name, start_column_letter,start_date,end_date)
    write_weeks_as_headers(destination_wb,sheet_name, start_column_letter,start_date,end_date)
    
    # Caculate range_format
    end_column_index = openpyxl.utils.column_index_from_string(start_column_letter) + date_count -1
    end_column_letter = get_column_letter(end_column_index)

    apply_conditional_formatting(destination_wb, sheet_name, 
                                 f'{start_column_letter}{EXCEL_FIRST_EMPLOYEE_ROW}:{end_column_letter}{EXCEL_LAST_EMPLOYEE_ROW}')

    # Lock the worksheet
    workbook = openpyxl.load_workbook(destination_wb)
    # Select the worksheet
    worksheet = workbook[sheet_name]
    # Protect the sheet with a password
    worksheet.protection.sheet = True
    worksheet.protection.password = CUSTOM_PASSWORD
    workbook.save(destination_wb)

def main():
    source_wb='vaccation-template.xlsx'
    destination_wb='2024-vaccation.xlsx'
    create_vaccation_period(source_wb=source_wb, destination_wb=destination_wb, sheet_name="Test-January-April",
                             start_date=date(CUSTOM_YEAR,1,1), end_date=date(CUSTOM_YEAR,4,30))
    create_vaccation_period(source_wb=destination_wb, destination_wb=destination_wb, sheet_name="Test-May-August",
                             start_date=date(CUSTOM_YEAR,5,1), end_date=date(CUSTOM_YEAR,8,31))
    create_vaccation_period(source_wb=destination_wb, destination_wb=destination_wb, sheet_name="Test-September-December",
                             start_date=date(CUSTOM_YEAR,9,1), end_date=date(CUSTOM_YEAR,12,31))

def create_textbox_and_connector(wb, ws):
  """Creates a textbox and a connector in the specified worksheet.

  Args:
    wb: The Excel workbook object.
    ws: The worksheet object.
  """

  # Create a new Shape object for the textbox.
  textbox = openpyxl.drawing.shapes.Shape()

  # Set the textbox's position and size.
  textbox.left = 10
  textbox.top = 10
  textbox.width = 100
  textbox.height = 20

  # Add text to the textbox.
  textbox.text = "This is a textbox."

  # Create a new Shape object for the connector.
  connector = openpyxl.drawing.shapes.Shape()

  # Set the connector's start and end points.
  connector.start = textbox.upperLeftCorner
  connector.end = (textbox.left + textbox.width, textbox.top + textbox.height)

  # Set the connector's style.
  connector.lineStyle = openpyxl.styles.LineFormat()
  connector.lineStyle.color = openpyxl.styles.Color(rgb="FF0000")
  connector.lineStyle.width = 1

  # Add the connector to the textbox.
  textbox.add_shape(connector)

if __name__ == "__main__":
    main()

